import csv
import io
import json
from pathlib import Path
from urllib.parse import quote

import xmltodict
from pydantic import BaseModel, Field

from configuration import get_session, get_system_config
from connection.connection import ensure_login
from generics import ApiResponse, FileTransferOutput, FileTransferResponse


DATAPREVIEW_ACCEPT = "application/xml, application/vnd.sap.adt.datapreview.table.v1+xml"
DATAPREVIEW_TABLE_ACCEPT = "application/vnd.sap.adt.datapreview.table.v1+xml"
SUPPORTED_OUTPUT_FORMATS = {"raw", "md", "csv"}
SUPPORTED_FILE_FORMATS = {"raw", "md", "csv", "xlsx"}


class DataPreviewColumnMetadata(BaseModel):
    """Column metadata parsed from SAP ADT data preview XML, useful for understanding what each returned value means."""

    name: str = Field(..., description="Technical column name as returned by SAP, e.g. BNAME or MANDT. Use this key to access values in rows.")
    type: str = Field("", description="SAP data preview primitive type, e.g. C character, N numeric text, D date, T time, P packed number, I integer or F float.")
    description: str = Field("", description="Human-readable DDIC field description when SAP provides one; useful for explaining technical column names.")
    colType: str = Field("", description="DDIC column type such as CHAR, CLNT, NUMC, DATS or TIMS when SAP provides it.")
    length: int | None = Field(None, description="Field length reported by SAP/DDIC. For character fields this is the maximum string length.")
    keyAttribute: bool = Field(False, description="True when SAP marks the field as part of the preview key metadata.")
    isKeyFigure: bool = Field(False, description="True when SAP marks the field as a key figure or measure.")
    caseSensitive: bool | None = Field(None, description="Whether SAP says comparisons for this field are case-sensitive; null means SAP did not report it.")


class DataPreviewMetadataOutput(BaseModel):
    """Field catalog for a DDIC/freestyle data preview result."""

    entity: str = Field(..., description="DDIC entity name for DDIC preview results, or empty for freestyle queries when SAP does not return an entity name.")
    totalRows: int | None = Field(None, description="Total rows value reported by SAP. SAP often returns 0 even when data rows are included, so prefer rowCount for parsed result size.")
    isHanaAnalyticalView: bool | None = Field(None, description="Whether SAP marks the previewed object as a HANA analytical view.")
    columns: list[DataPreviewColumnMetadata] = Field(default_factory=list, description="Ordered field catalog. This order matches CSV, Markdown, XLSX and row object columns.")


class DataPreviewResultOutput(BaseModel):
    """Data preview query result formatted for AI consumption."""

    format: str = Field(..., description="Returned inline format: raw, md or csv. XLSX is only available in the *_to_file tools.")
    content: str = Field(..., description="Formatted result payload. For raw this is SAP XML; for md/csv this is a row-oriented table converted from SAP's column-oriented XML.")
    metadata: DataPreviewMetadataOutput | None = Field(None, description="Field catalog for md/csv outputs. Null for raw because raw already contains the original SAP XML.")
    rowCount: int = Field(0, description="Number of data rows parsed from the ADT response after transposing SAP's column-oriented XML.")
    query: str = Field("", description="ABAP Open SQL query sent to SAP. For DDIC preview this may be generated from metadata plus the optional where condition.")
    executedQuery: str = Field("", description="Query string returned by SAP, mainly for freestyle queries where SAP expands the ABAP Open SQL statement.")
    queryExecutionTime: str = Field("", description="Execution time reported by SAP as text, when available.")


class DataPreviewFileOutput(FileTransferOutput):
    """Local file export result for data preview outputs."""

    metadataFilePath: str | None = Field(None, description="Local path of the JSON sidecar metadata file. Null for raw and xlsx exports because xlsx embeds metadata in a worksheet.")
    metadataSizeBytes: int | None = Field(None, description="Number of bytes written to the JSON sidecar metadata file, when one is created.")


class DataPreviewMetadataResponse(ApiResponse[DataPreviewMetadataOutput]):
    """Response model for DDIC data preview metadata."""


class DataPreviewResultResponse(ApiResponse[DataPreviewResultOutput]):
    """Response model for DDIC/freestyle data preview queries."""


class DataPreviewFileResponse(ApiResponse[DataPreviewFileOutput]):
    """Response model for DDIC/freestyle data preview exports."""


def _bool_or_none(value) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    lowered = str(value).strip().lower()
    if lowered in {"true", "x", "1"}:
        return True
    if lowered in {"false", "", "0"}:
        return False
    return None


def _int_or_none(value) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _as_list(value) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _metadata_from_node(node: dict) -> DataPreviewColumnMetadata:
    return DataPreviewColumnMetadata(
        name=node.get("@dataPreview:name") or node.get("@name") or "",
        type=node.get("@dataPreview:type") or node.get("@type") or "",
        description=node.get("@dataPreview:description") or node.get("@description") or "",
        colType=node.get("@dataPreview:colType") or node.get("@colType") or "",
        length=_int_or_none(node.get("@dataPreview:length") or node.get("@length")),
        keyAttribute=bool(_bool_or_none(node.get("@dataPreview:keyAttribute") or node.get("@keyAttribute"))),
        isKeyFigure=bool(_bool_or_none(node.get("@dataPreview:isKeyFigure") or node.get("@isKeyFigure"))),
        caseSensitive=_bool_or_none(node.get("@dataPreview:caseSensitive") or node.get("@caseSensitive")),
    )


def _parse_table_data_xml(xml_text: str, default_entity: str = "") -> tuple[DataPreviewMetadataOutput, list[dict[str, str]], dict[str, str]]:
    parsed = xmltodict.parse(xml_text)
    table = parsed.get("dataPreview:tableData") or parsed.get("tableData") or {}
    entity = table.get("dataPreview:name") or table.get("name") or default_entity
    columns = []
    column_values = []

    for column_node in _as_list(table.get("dataPreview:columns") or table.get("columns")):
        metadata_node = column_node.get("dataPreview:metadata") or column_node.get("metadata") or {}
        column_metadata = _metadata_from_node(metadata_node)
        columns.append(column_metadata)

        data_set = column_node.get("dataPreview:dataSet") or column_node.get("dataSet") or {}
        values = _as_list(data_set.get("dataPreview:data") or data_set.get("data"))
        normalized_values = []
        for value in values:
            if isinstance(value, dict):
                normalized_values.append(str(value.get("#text") or ""))
            elif value is None:
                normalized_values.append("")
            else:
                normalized_values.append(str(value))
        column_values.append(normalized_values)

    row_count = max((len(values) for values in column_values), default=0)
    rows: list[dict[str, str]] = []
    for row_index in range(row_count):
        row = {}
        for column, values in zip(columns, column_values):
            row[column.name] = values[row_index] if row_index < len(values) else ""
        rows.append(row)

    metadata = DataPreviewMetadataOutput(
        entity=entity,
        totalRows=_int_or_none(table.get("dataPreview:totalRows") or table.get("totalRows")),
        isHanaAnalyticalView=_bool_or_none(table.get("dataPreview:isHanaAnalyticalView") or table.get("isHanaAnalyticalView")),
        columns=columns,
    )
    extras = {
        "executedQuery": table.get("dataPreview:executedQueryString") or table.get("executedQueryString") or "",
        "queryExecutionTime": table.get("dataPreview:queryExecutionTime") or table.get("queryExecutionTime") or "",
    }
    return metadata, rows, extras


def _normalize_output_format(outputFormat: str) -> str:
    normalized = (outputFormat or "md").strip().lower()
    if normalized not in SUPPORTED_OUTPUT_FORMATS:
        raise ValueError("outputFormat must be one of: raw, md, csv.")
    return normalized


def _normalize_file_format(outputFormat: str) -> str:
    normalized = (outputFormat or "csv").strip().lower()
    if normalized not in SUPPORTED_FILE_FORMATS:
        raise ValueError("outputFormat must be one of: raw, md, csv, xlsx.")
    return normalized


def _rows_to_csv(columns: list[DataPreviewColumnMetadata], rows: list[dict[str, str]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[column.name for column in columns], lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def _escape_md_cell(value: str) -> str:
    return str(value).replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def _rows_to_markdown(columns: list[DataPreviewColumnMetadata], rows: list[dict[str, str]]) -> str:
    headers = [column.name for column in columns]
    if not headers:
        return ""
    lines = [
        "| " + " | ".join(_escape_md_cell(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_escape_md_cell(row.get(header, "")) for header in headers) + " |")
    return "\n".join(lines)


def _format_content(output_format: str, raw_xml: str, metadata: DataPreviewMetadataOutput, rows: list[dict[str, str]]) -> str:
    if output_format == "raw":
        return raw_xml
    if output_format == "csv":
        return _rows_to_csv(metadata.columns, rows)
    return _rows_to_markdown(metadata.columns, rows)


def _metadata_to_json(metadata: DataPreviewMetadataOutput) -> str:
    return metadata.model_dump_json(indent=2)


def _metadata_payload(result: DataPreviewResultOutput) -> dict:
    return {
        "metadata": result.metadata.model_dump(mode="json") if result.metadata else None,
        "query": result.query,
        "executedQuery": result.executedQuery,
        "queryExecutionTime": result.queryExecutionTime,
        "rowCount": result.rowCount,
    }


def _resolve_output_path(filePath: str) -> Path:
    path = Path(filePath)
    if not path.is_absolute():
        path = Path.cwd() / path
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _write_text(path: Path, content: str) -> int:
    path.write_text(content, encoding="utf-8", newline="\n")
    return len(content.encode("utf-8"))


def _write_xlsx(path: Path, result: DataPreviewResultOutput, rows: list[dict[str, str]]) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if result.metadata is None:
        raise ValueError("XLSX output requires parsed metadata and cannot be generated from raw output.")

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "data"
    headers = [column.name for column in result.metadata.columns]
    data_sheet.append(headers)
    for cell in data_sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        data_sheet.append([row.get(header, "") for header in headers])
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = data_sheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(row.get(header, "")) for row in rows[:100]]
        max_width = max([len(str(header)), *(len(value) for value in values)], default=10)
        data_sheet.column_dimensions[get_column_letter(index)].width = min(max(max_width + 2, 10), 60)

    metadata_sheet = workbook.create_sheet("metadata")
    metadata_headers = ["name", "type", "colType", "length", "description", "keyAttribute", "isKeyFigure", "caseSensitive"]
    metadata_sheet.append(metadata_headers)
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True)
    for column in result.metadata.columns:
        metadata_sheet.append([
            column.name,
            column.type,
            column.colType,
            column.length,
            column.description,
            column.keyAttribute,
            column.isKeyFigure,
            column.caseSensitive,
        ])
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.auto_filter.ref = metadata_sheet.dimensions
    for index, header in enumerate(metadata_headers, start=1):
        metadata_sheet.column_dimensions[get_column_letter(index)].width = 18 if header != "description" else 45

    query_sheet = workbook.create_sheet("query")
    query_sheet.append(["property", "value"])
    for cell in query_sheet[1]:
        cell.font = Font(bold=True)
    for key, value in _metadata_payload(result).items():
        if key == "metadata":
            continue
        query_sheet.append([key, value])
    query_sheet.column_dimensions["A"].width = 24
    query_sheet.column_dimensions["B"].width = 100

    workbook.save(path)
    return path.stat().st_size


def _metadata_sidecar_path(path: Path) -> Path:
    return Path(f"{path}.metadata")


def _base_url(systemId: str, path: str) -> str:
    system_config = get_system_config(systemId)
    return f"{system_config.server}{path}"


def _session_get(systemId: str, path: str, *, headers: dict[str, str], params: dict[str, str] | None = None):
    return get_session(systemId).get(_base_url(systemId, path), headers=headers, params=params)


def _session_post(systemId: str, path: str, *, headers: dict[str, str], params: dict[str, str], data: str):
    return get_session(systemId).post(_base_url(systemId, path), headers=headers, params=params, data=data.encode("utf-8"))


def _login_error_response(response_type, action: str, error_msg: str):
    return response_type.model_validate({
        "result": False,
        "httpCode": 401,
        "httpReason": "Unauthorized",
        "message": f"Cannot {action} because no SAP session is available: {error_msg}",
        "data": None,
    })


def _http_error_response(response_type, response, action: str):
    return response_type.model_validate({
        "result": False,
        "httpCode": response.status_code,
        "httpReason": response.reason,
        "message": f"ADT rejected the {action} request: {response.text}",
        "data": None,
    })


def _get_metadata(systemId: str, ddicEntityName: str):
    entity = ddicEntityName.upper()
    return _session_get(
        systemId,
        f"/sap/bc/adt/datapreview/ddic/{quote(entity, safe='')}/metadata",
        headers={"Accept": DATAPREVIEW_TABLE_ACCEPT},
    )


def _post_ddic_query(systemId: str, ddicEntityName: str, rowNumber: int, sql_query: str):
    return _session_post(
        systemId,
        "/sap/bc/adt/datapreview/ddic",
        headers={
            "Accept": DATAPREVIEW_ACCEPT,
            "Content-Type": "text/plain",
        },
        params={"rowNumber": str(rowNumber), "ddicEntityName": ddicEntityName.upper()},
        data=sql_query,
    )


def _post_freestyle_query(systemId: str, rowNumber: int, sql_query: str):
    return _session_post(
        systemId,
        "/sap/bc/adt/datapreview/freestyle",
        headers={
            "Accept": DATAPREVIEW_ACCEPT,
            "Content-Type": "text/plain",
        },
        params={"rowNumber": str(rowNumber)},
        data=sql_query,
    )


def _build_ddic_select(ddicEntityName: str, metadata: DataPreviewMetadataOutput, where: str = "") -> str:
    entity = ddicEntityName.upper()
    columns = [f"{entity}~{column.name}" for column in metadata.columns if column.name]
    if not columns:
        columns = ["*"]
    query = f"SELECT {', '.join(columns)} FROM {entity}"
    if where.strip():
        query = f"{query} WHERE {where.strip()}"
    return query


def call_datapreview_metadata(systemId: str, ddicEntityName: str) -> DataPreviewMetadataResponse:
    """Fetch metadata for one DDIC entity through ADT data preview."""
    try:
        is_logged_in, error_msg = ensure_login(systemId)
        if not is_logged_in:
            return _login_error_response(DataPreviewMetadataResponse, "read data preview metadata", error_msg)

        response = _get_metadata(systemId, ddicEntityName)
        if response.status_code != 200:
            return _http_error_response(DataPreviewMetadataResponse, response, "data preview metadata")

        metadata, _, _ = _parse_table_data_xml(response.text, ddicEntityName.upper())
        return DataPreviewMetadataResponse.model_validate({
            "result": True,
            "httpCode": response.status_code,
            "httpReason": response.reason,
            "message": "Data preview metadata read successfully.",
            "data": metadata,
        })
    except Exception as exc:
        return DataPreviewMetadataResponse.model_validate({
            "result": False,
            "httpCode": 500,
            "httpReason": "Internal Server Error",
            "message": f"Unexpected error while reading data preview metadata: {str(exc)}",
            "data": None,
        })


def call_datapreview_table_contents(
    systemId: str,
    ddicEntityName: str,
    rowNumber: int = 100,
    where: str = "",
    sqlQuery: str = "",
    outputFormat: str = "md",
) -> DataPreviewResultResponse:
    """Read DDIC entity contents through ADT data preview."""
    try:
        output_format = _normalize_output_format(outputFormat)
        is_logged_in, error_msg = ensure_login(systemId)
        if not is_logged_in:
            return _login_error_response(DataPreviewResultResponse, "read DDIC data preview contents", error_msg)

        metadata_response = _get_metadata(systemId, ddicEntityName)
        if metadata_response.status_code != 200:
            return _http_error_response(DataPreviewResultResponse, metadata_response, "data preview metadata")
        metadata, _, _ = _parse_table_data_xml(metadata_response.text, ddicEntityName.upper())

        query = sqlQuery.strip() or _build_ddic_select(ddicEntityName, metadata, where)
        response = _post_ddic_query(systemId, ddicEntityName, rowNumber, query)
        if response.status_code != 200:
            return _http_error_response(DataPreviewResultResponse, response, "DDIC data preview contents")

        parsed_metadata, rows, extras = _parse_table_data_xml(response.text, ddicEntityName.upper())
        content = _format_content(output_format, response.text, parsed_metadata, rows)
        return DataPreviewResultResponse.model_validate({
            "result": True,
            "httpCode": response.status_code,
            "httpReason": response.reason,
            "message": "DDIC data preview contents read successfully.",
            "data": DataPreviewResultOutput(
                format=output_format,
                content=content,
                metadata=None if output_format == "raw" else parsed_metadata,
                rowCount=len(rows),
                query=query,
                executedQuery=extras["executedQuery"],
                queryExecutionTime=extras["queryExecutionTime"],
            ),
        })
    except ValueError as exc:
        return DataPreviewResultResponse.model_validate({
            "result": False,
            "httpCode": 400,
            "httpReason": "Bad Request",
            "message": str(exc),
            "data": None,
        })
    except Exception as exc:
        return DataPreviewResultResponse.model_validate({
            "result": False,
            "httpCode": 500,
            "httpReason": "Internal Server Error",
            "message": f"Unexpected error while reading DDIC data preview contents: {str(exc)}",
            "data": None,
        })


def call_datapreview_run_query(
    systemId: str,
    sqlQuery: str,
    rowNumber: int = 100,
    outputFormat: str = "md",
) -> DataPreviewResultResponse:
    """Run a freestyle ABAP Open SQL query through ADT data preview."""
    try:
        output_format = _normalize_output_format(outputFormat)
        is_logged_in, error_msg = ensure_login(systemId)
        if not is_logged_in:
            return _login_error_response(DataPreviewResultResponse, "run freestyle data preview query", error_msg)
        if not sqlQuery.strip():
            raise ValueError("sqlQuery is required.")

        response = _post_freestyle_query(systemId, rowNumber, sqlQuery.strip())
        if response.status_code != 200:
            return _http_error_response(DataPreviewResultResponse, response, "freestyle data preview query")

        metadata, rows, extras = _parse_table_data_xml(response.text)
        content = _format_content(output_format, response.text, metadata, rows)
        return DataPreviewResultResponse.model_validate({
            "result": True,
            "httpCode": response.status_code,
            "httpReason": response.reason,
            "message": "Freestyle ABAP Open SQL data preview query executed successfully.",
            "data": DataPreviewResultOutput(
                format=output_format,
                content=content,
                metadata=None if output_format == "raw" else metadata,
                rowCount=len(rows),
                query=sqlQuery.strip(),
                executedQuery=extras["executedQuery"],
                queryExecutionTime=extras["queryExecutionTime"],
            ),
        })
    except ValueError as exc:
        return DataPreviewResultResponse.model_validate({
            "result": False,
            "httpCode": 400,
            "httpReason": "Bad Request",
            "message": str(exc),
            "data": None,
        })
    except Exception as exc:
        return DataPreviewResultResponse.model_validate({
            "result": False,
            "httpCode": 500,
            "httpReason": "Internal Server Error",
            "message": f"Unexpected error while running freestyle data preview query: {str(exc)}",
            "data": None,
        })


def _rows_from_csv_content(content: str) -> list[dict[str, str]]:
    return [dict(row) for row in csv.DictReader(io.StringIO(content))]


def _write_result_to_file(filePath: str, result: DataPreviewResultOutput, uri: str) -> DataPreviewFileResponse:
    output_path = _resolve_output_path(filePath)
    if result.format == "xlsx":
        size_bytes = _write_xlsx(output_path, result, _rows_from_csv_content(result.content))
    else:
        size_bytes = _write_text(output_path, result.content)
    metadata_path = None
    metadata_size = None

    if result.metadata is not None and result.format != "xlsx":
        metadata_path = _metadata_sidecar_path(output_path)
        metadata_size = _write_text(metadata_path, json.dumps(_metadata_payload(result), indent=2, ensure_ascii=False))

    mime_type = {
        "raw": "application/vnd.sap.adt.datapreview.table.v1+xml",
        "csv": "text/csv",
        "md": "text/markdown",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }.get(result.format, "text/plain")

    return DataPreviewFileResponse.model_validate({
        "result": True,
        "httpCode": 200,
        "httpReason": "OK",
        "message": "Data preview result written to local file successfully.",
        "data": DataPreviewFileOutput(
            filePath=str(output_path),
            uri=uri,
            mimeType=mime_type,
            sizeBytes=size_bytes,
            metadataFilePath=str(metadata_path) if metadata_path else None,
            metadataSizeBytes=metadata_size,
        ),
    })


def call_datapreview_table_contents_to_file(
    systemId: str,
    ddicEntityName: str,
    filePath: str,
    rowNumber: int = 100,
    where: str = "",
    sqlQuery: str = "",
    outputFormat: str = "csv",
) -> DataPreviewFileResponse:
    """Read DDIC entity contents and write the converted data to a local file."""
    try:
        output_format = _normalize_file_format(outputFormat)
        query_format = "csv" if output_format == "xlsx" else output_format
        response = call_datapreview_table_contents(systemId, ddicEntityName, rowNumber, where, sqlQuery, query_format)
        if not response.result or response.data is None:
            return DataPreviewFileResponse.model_validate(response.model_dump())
        if output_format == "xlsx":
            response.data.format = "xlsx"
        return _write_result_to_file(filePath, response.data, f"/sap/bc/adt/datapreview/ddic?ddicEntityName={ddicEntityName.upper()}")
    except ValueError as exc:
        return DataPreviewFileResponse.model_validate({
            "result": False,
            "httpCode": 400,
            "httpReason": "Bad Request",
            "message": str(exc),
            "data": None,
        })
    except Exception as exc:
        return DataPreviewFileResponse.model_validate({
            "result": False,
            "httpCode": 500,
            "httpReason": "Internal Server Error",
            "message": f"Failed to write DDIC data preview result to file: {str(exc)}",
            "data": None,
        })


def call_datapreview_run_query_to_file(
    systemId: str,
    sqlQuery: str,
    filePath: str,
    rowNumber: int = 100,
    outputFormat: str = "csv",
) -> DataPreviewFileResponse:
    """Run a freestyle ABAP Open SQL query and write the converted data to a local file."""
    try:
        output_format = _normalize_file_format(outputFormat)
        query_format = "csv" if output_format == "xlsx" else output_format
        response = call_datapreview_run_query(systemId, sqlQuery, rowNumber, query_format)
        if not response.result or response.data is None:
            return DataPreviewFileResponse.model_validate(response.model_dump())
        if output_format == "xlsx":
            response.data.format = "xlsx"
        return _write_result_to_file(filePath, response.data, "/sap/bc/adt/datapreview/freestyle")
    except ValueError as exc:
        return DataPreviewFileResponse.model_validate({
            "result": False,
            "httpCode": 400,
            "httpReason": "Bad Request",
            "message": str(exc),
            "data": None,
        })
    except Exception as exc:
        return DataPreviewFileResponse.model_validate({
            "result": False,
            "httpCode": 500,
            "httpReason": "Internal Server Error",
            "message": f"Failed to write freestyle data preview result to file: {str(exc)}",
            "data": None,
        })
